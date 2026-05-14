"""
================================================================================
QM 자동화 관제시스템
================================================================================
Convert Pro 3 에 내장된 실시간 모니터링 웹 대시보드.
변환 완료된 CSV 데이터를 브라우저에서 바로 확인할 수 있도록 Flask 로 제공한다.

주요 기능
---------
- 변환본 CSV 파일의 최신값·추이 실시간 조회
- 장비·현장별 상태 현황판
- Convert_pro3.py 앱과 동일 프로세스 내 daemon 스레드로 동작
- 로그인 세션 기반 접근 제어

실행 방법
---------
  직접 실행 : python monitoring/server.py   (브라우저 자동 열림)
  앱 통합   : Convert_pro3.py 에서 start_server() 를 daemon thread 로 호출

프로젝트 내 위치
----------------
  Convert_pro3.py          컨버트 프로그램 (진입점)
  monitoring/server.py     ← 이 파일 (QM 자동화 관제시스템)
  measurement_portal/      계측관리 통합시스템
================================================================================
"""
import json
import os
import sqlite3
import sys
import threading
import time
from io import BytesIO
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import secrets
from functools import wraps

from flask import Flask, jsonify, render_template, request, session, redirect, url_for, send_file

# ─────────────────────────────────────────
#  경로 설정
# ─────────────────────────────────────────
if getattr(sys, "frozen", False):
    _meipass = Path(sys._MEIPASS)
    _this_file = Path(__file__).resolve()
    if str(_this_file).startswith(str(_meipass)):
        # PyInstaller 번들 내 내장 파일 (폴백)
        _APP_ROOT = Path(sys.executable).parent
        _TEMPLATE_FOLDER = _meipass / "monitoring" / "templates"
    else:
        # exe 옆에 외부 파일로 존재 (편집 가능 모드)
        _APP_ROOT = _this_file.parent.parent
        _TEMPLATE_FOLDER = _this_file.parent / "templates"
else:
    _APP_ROOT = Path(__file__).parent.parent
    _TEMPLATE_FOLDER = Path(__file__).parent / "templates"

CONFIG_PATH = _APP_ROOT / "config.json"
MGMT_PATH   = _APP_ROOT / "management.json"
AUTH_PATH   = _APP_ROOT / "web_auth.json"
QM_REMOTE_PATH = _APP_ROOT / "qm_remote.json"
QM_LOCAL_DB_PATH = _APP_ROOT / "qm_remote_state.sqlite3"
_QM_DB_LOCK = threading.Lock()
WEB_SHORTCUTS_PATH = _APP_ROOT / "web_shortcuts.json"
CONVERT_ROOT = r"C:\data\Convertfile"
PORT = 5050

# CH0~CH7 = CSV 컬럼 인덱스 16~23 고정
CH_COL_START = 16

# ─────────────────────────────────────────
#  인증 설정 (web_auth.json)
# ─────────────────────────────────────────
def _load_auth_cfg() -> dict:
    """web_auth.json 읽기. 없으면 기본값 생성."""
    if AUTH_PATH.exists():
        try:
            with open(AUTH_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    # 파일 없음 → 기본 자격증명 + 랜덤 secret_key 생성
    cfg = {
        "username": "admin",
        "password": "changeme",
        "secret_key": secrets.token_hex(32),
    }
    try:
        with open(AUTH_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        print(f"[Auth] web_auth.json 생성됨 → {AUTH_PATH}", flush=True)
        print(f"[Auth] 초기 아이디: admin  /  비밀번호: changeme  (꼭 변경하세요!)", flush=True)
    except Exception:
        pass
    return cfg

_auth_cfg = _load_auth_cfg()

app = Flask(__name__, template_folder=str(_TEMPLATE_FOLDER))
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.secret_key = _auth_cfg.get("secret_key") or secrets.token_hex(32)

# ─────────────────────────────────────────
#  인증 데코레이터
# ─────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": "인증 필요", "auth_required": True}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated


_LOGIN_HTML = """<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>큐엠 자동화 관리 프로그램 — 로그인</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{min-height:100vh;display:flex;align-items:center;justify-content:center;
       background:#f1f5f9;font-family:'Segoe UI',sans-serif}
  .card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;
        padding:40px 36px;width:340px;box-shadow:0 4px 24px rgba(0,0,0,.08)}
  .logo{font-size:22px;font-weight:800;color:#1e40af;margin-bottom:4px}
  .sub{font-size:13px;color:#64748b;margin-bottom:28px}
  label{display:block;font-size:12px;font-weight:600;color:#475569;
        text-transform:uppercase;letter-spacing:.4px;margin-bottom:6px}
  input{width:100%;padding:10px 12px;border:1px solid #cbd5e1;border-radius:8px;
        font-size:14px;color:#1e293b;outline:none;transition:border .15s;margin-bottom:16px}
  input:focus{border-color:#3b82f6;box-shadow:0 0 0 3px rgba(59,130,246,.15)}
  button{width:100%;padding:11px;background:#1d4ed8;color:#fff;border:none;
         border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;transition:background .15s;margin-top:4px}
  button:hover{background:#1e40af}
  .err{color:#dc2626;font-size:13px;margin-bottom:14px;background:#fef2f2;
       border:1px solid #fecaca;border-radius:6px;padding:8px 12px}
</style>
</head>
<body>
<div class="card">
  <div class="logo">큐엠 자동화 관리 프로그램</div>
  <div class="sub">현장관리 웹 대시보드</div>
  {error}
  <form method="post">
    <label>아이디</label>
    <input name="username" type="text" autocomplete="username" autofocus>
    <label>비밀번호</label>
    <input name="password" type="password" autocomplete="current-password">
    <button type="submit">로그인</button>
  </form>
</div>
</body>
</html>"""


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("logged_in"):
        return redirect("/")
    error_html = ""
    if request.method == "POST":
        u = request.form.get("username", "")
        p = request.form.get("password", "")
        cfg = _load_auth_cfg()   # 항상 최신 자격증명 읽기
        if u == cfg.get("username") and p == cfg.get("password"):
            session["logged_in"] = True
            session.permanent = False
            return redirect("/")
        error_html = '<div class="err">아이디 또는 비밀번호가 올바르지 않습니다</div>'
    return _LOGIN_HTML.replace("{error}", error_html), 200, {"Content-Type": "text/html; charset=utf-8"}


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


def _qm_remote_anonymous_token_authorized() -> bool:
    """데스크톱·자동 클라이언트: X-QM-Remote-Token + qm_remote.json desktop_api_token 일치 시 QM API만 통과."""
    if not request.path.startswith("/api/qm-remote/"):
        return False
    try:
        cfg = _load_qm_remote_config()
    except Exception:
        return False
    tok = str(cfg.get("desktop_api_token") or "").strip()
    if not tok:
        return False
    got = (request.headers.get("X-QM-Remote-Token") or "").strip()
    return bool(got) and secrets.compare_digest(got, tok)


@app.before_request
def _require_login():
    """로그인/로그아웃·static 을 제외한 모든 요청에 인증 적용."""
    if request.endpoint in ("login", "logout", "static"):
        return
    if _qm_remote_anonymous_token_authorized():
        return
    if not session.get("logged_in"):
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": "인증 필요", "auth_required": True}), 401
        return redirect("/login")


# ─────────────────────────────────────────
#  데스크탑 앱 트리 새로고침 콜백
#  Convert_pro3.py 에서 set_category_saved_callback() 으로 등록
# ─────────────────────────────────────────
_on_category_saved = None

def set_category_saved_callback(fn):
    """카테고리 저장 완료 시 호출할 함수를 등록한다.
    fn()은 Tkinter 메인 루프에서 스케줄링하는 래퍼여야 한다.
    """
    global _on_category_saved
    _on_category_saved = fn


# ─────────────────────────────────────────
#  config.json 읽기 / 쓰기
#  ※ _CONFIG_LOCK 은 ConfigManager.save_lock 과 별개이지만
#    server.py 내에서 read-modify-write 전 구간을 하나의 락으로 직렬화한다.
#    원자적 쓰기(temp → os.replace)로 부분 쓰기 노출을 방지한다.
# ─────────────────────────────────────────

_CONFIG_LOCK = threading.Lock()
_MGMT_LOCK   = threading.Lock()

_EMPTY_SITE_MGMT = {
    "address": "", "program": "", "memo": "",
    "check_interval": 0, "report_enabled": False, "report_cycle": "",
    "assigned_manager_id": "", "assigned_qm_manager_id": "",
    "station_groups": [], "stations": [], "assignments": {}
}

_DEFAULT_PROGRAM_TYPES = ["AMS 휴먼프로그램", "새길 계측 프로그램"]


def _get_program_types_list(mgmt: dict) -> list:
    """management.json 의 program_types 또는 기본 목록."""
    raw = mgmt.get("program_types")
    if isinstance(raw, list):
        seen = set()
        out = []
        for x in raw:
            s = str(x).strip()
            if s and s not in seen:
                seen.add(s)
                out.append(s)
        if out:
            return out
    return list(_DEFAULT_PROGRAM_TYPES)


def _load_config() -> dict:
    try:
        with open(CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_config(cfg: dict):
    """원자적 저장: 락 획득 → 임시 파일 기록 → os.replace() 교체.

    - 락(_CONFIG_LOCK)으로 동시 저장 직렬화
    - temp→rename 방식으로 부분 쓰기 노출 방지
    - ConfigManager.save_lock 과는 별개이나, 원자 쓰기 덕분에 파일 깨짐 없음
    """
    with _CONFIG_LOCK:
        tmp = Path(str(CONFIG_PATH) + ".tmp")
        try:
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=4)
            os.replace(tmp, CONFIG_PATH)
        except Exception:
            if tmp.exists():
                tmp.unlink(missing_ok=True)
            raise


@contextmanager
def _edit_config():
    """config.json read-modify-write 직렬화 컨텍스트 매니저."""
    with _CONFIG_LOCK:
        cfg = _load_config()
        try:
            yield cfg
        except Exception:
            raise
        else:
            tmp = Path(str(CONFIG_PATH) + ".tmp")
            try:
                with open(tmp, "w", encoding="utf-8") as f:
                    json.dump(cfg, f, ensure_ascii=False, indent=4)
                os.replace(tmp, CONFIG_PATH)
            except Exception:
                if tmp.exists():
                    tmp.unlink(missing_ok=True)
                raise


# ─────────────────────────────────────────
#  management.json 읽기 / 쓰기
# ─────────────────────────────────────────

def _load_management() -> dict:
    try:
        with open(MGMT_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"__version__": 1, "qm_managers": [], "companies": {}}


def _save_management(mgmt: dict):
    """원자적 저장."""
    with _MGMT_LOCK:
        tmp = Path(str(MGMT_PATH) + ".tmp")
        try:
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(mgmt, f, ensure_ascii=False, indent=4)
            os.replace(tmp, MGMT_PATH)
        except Exception:
            if tmp.exists():
                tmp.unlink(missing_ok=True)
            raise


@contextmanager
def _edit_management():
    """management.json read-modify-write 직렬화 컨텍스트 매니저."""
    with _MGMT_LOCK:
        mgmt = _load_management()
        try:
            yield mgmt
        except Exception:
            raise
        else:
            tmp = Path(str(MGMT_PATH) + ".tmp")
            try:
                with open(tmp, "w", encoding="utf-8") as f:
                    json.dump(mgmt, f, ensure_ascii=False, indent=4)
                os.replace(tmp, MGMT_PATH)
            except Exception:
                if tmp.exists():
                    tmp.unlink(missing_ok=True)
                raise


def _find_site_mgmt(mgmt: dict, company: str, site: str) -> dict | None:
    """management.json에서 현장 데이터 반환. 없으면 None."""
    return (
        mgmt.get("companies", {})
            .get(company, {})
            .get("sites", {})
            .get(site)
    )


def _ensure_site_mgmt(mgmt: dict, company: str, site: str) -> dict:
    """management.json에서 현장 노드를 보장하며 반환."""
    comp = mgmt.setdefault("companies", {}).setdefault(
        company, {"managers": [], "sites": {}}
    )
    sites = comp.setdefault("sites", {})
    if site not in sites:
        import copy
        sites[site] = copy.deepcopy(_EMPTY_SITE_MGMT)
    return sites[site]


# ─────────────────────────────────────────
#  CSV 마지막 N행 파싱
# ─────────────────────────────────────────

def _read_csv_tail(out_path: str, n: int = 48):
    """
    출력 CSV 마지막 n행 파싱.
    Returns: (timestamps: list[str], ch_values: dict[str, list[float|None]])
    """
    empty = {f"CH{i}": [] for i in range(8)}
    if not os.path.exists(out_path):
        return [], empty
    try:
        with open(out_path, "rb") as f:
            f.seek(0, 2)
            size = f.tell()
            f.seek(max(0, size - 1024 * 48))
            tail = f.read().decode("utf-8-sig", errors="replace")

        data_lines = []
        for line in tail.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            low = line.lower().lstrip('"')
            if low.startswith("timestamp"):
                continue
            data_lines.append(line)

        data_lines = data_lines[-n:]
        timestamps = []
        ch_values = {f"CH{i}": [] for i in range(8)}

        for line in data_lines:
            parts = [p.strip().strip('"') for p in line.split(",")]
            timestamps.append(parts[0] if parts else "")
            for i in range(8):
                idx = CH_COL_START + i
                try:
                    v = float(parts[idx]) if idx < len(parts) and parts[idx] else None
                except (ValueError, TypeError):
                    v = None
                ch_values[f"CH{i}"].append(v)

        return timestamps, ch_values
    except Exception:
        return [], empty


# ─────────────────────────────────────────
#  마지막 timestamp 읽기
# ─────────────────────────────────────────

def _last_timestamp(out_path: str):
    try:
        if not os.path.exists(out_path):
            return None
        with open(out_path, "rb") as f:
            f.seek(0, 2)
            size = f.tell()
            f.seek(max(0, size - 4096))
            tail = f.read().decode("utf-8-sig", errors="replace")
        lines = [ln.strip() for ln in tail.splitlines() if ln.strip()]
        for line in reversed(lines):
            if line.startswith("#") or line.lower().lstrip('"').startswith("timestamp"):
                continue
            parts = line.split(",")
            if parts:
                ts = parts[0].strip().strip('"')
                try:
                    return datetime.fromisoformat(ts)
                except Exception:
                    import re
                    m = re.match(r"(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2})", ts)
                    if m:
                        return datetime.fromisoformat(m.group(1))
        return None
    except Exception:
        return None


# ─────────────────────────────────────────
#  파일 상태 계산
# ─────────────────────────────────────────

def _file_status(out_path: str, is_ghost: bool) -> dict:
    if is_ghost:
        return {"status": "Ghost", "last_ts": "-", "elapsed": "-", "elapsed_h": 9999}
    if not os.path.exists(out_path):
        return {"status": "미변환", "last_ts": "-", "elapsed": "-", "elapsed_h": 9999}
    last_ts = _last_timestamp(out_path)
    if last_ts is None:
        return {"status": "미변환", "last_ts": "-", "elapsed": "-", "elapsed_h": 9999}
    now = datetime.now()
    delta = now - last_ts
    hours = delta.total_seconds() / 3600
    total_min = int(delta.total_seconds() / 60)
    h, m = divmod(total_min, 60)
    elapsed_str = f"{h}시간 {m}분" if h else f"{m}분"
    status = "정상" if hours < 24 else "지연"
    return {
        "status": status,
        "last_ts": last_ts.strftime("%Y-%m-%d %H:%M"),
        "elapsed": elapsed_str,
        "elapsed_h": round(hours, 2),
    }


# ─────────────────────────────────────────
#  채널 분석
# ─────────────────────────────────────────

def _parse_channels(file_cfg: dict) -> list:
    channels = []
    for i in range(8):
        key = f"CH{i}"
        ch = file_cfg.get(key, {})
        mode = (ch.get("offset") or "PASS").upper()
        label = ch.get("label", "").strip()
        active = bool(label) or (mode not in ("PASS", ""))
        channels.append({
            "ch":          key,
            "label":       label or "-",
            "mode":        mode or "PASS",
            "active":      active,
            "sensor_type": ch.get("sensor_type", ""),
        })
    return channels


def _fetch_external_485_tree(cfg: dict) -> list:
    """
    외부 485 서버의 파일 목록을 현재 트리 형식으로 변환.
    우선순위:
      1) 환경변수 EXTERNAL_485_LIST_URL
      2) config.json의 __external_485__.list_url
    """
    ext_cfg = cfg.get("__external_485__", {}) if isinstance(cfg, dict) else {}
    enabled = bool(ext_cfg.get("enabled", False))
    if not enabled:
        return []
    list_url = (
        (os.getenv("EXTERNAL_485_LIST_URL") or "").strip()
        or str(ext_cfg.get("list_url", "")).strip()
    )
    if not list_url:
        return []

    timeout_sec = 3
    try:
        timeout_sec = int(ext_cfg.get("timeout_sec", 3))
    except Exception:
        timeout_sec = 3

    try:
        req = Request(list_url, headers={"User-Agent": "ConvertPro3-Monitoring/1.0"})
        with urlopen(req, timeout=max(1, timeout_sec)) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except (HTTPError, URLError, TimeoutError, ValueError) as e:
        print(f"[MonitoringServer] external_485 fetch 실패: {e}", file=sys.stderr)
        return []
    except Exception as e:
        print(f"[MonitoringServer] external_485 예외: {e}", file=sys.stderr)
        return []

    # 케이스 A: 이미 api/tree 포맷 일부를 제공하는 경우
    if isinstance(payload, dict) and isinstance(payload.get("tree"), list):
        tree = payload.get("tree", [])
        # 최소 스키마 검증 + source 주입
        for comp in tree:
            if not isinstance(comp, dict):
                continue
            if "company" not in comp:
                comp["company"] = "485"
            if not isinstance(comp.get("sites"), list):
                comp["sites"] = []
            for site in comp.get("sites", []):
                if not isinstance(site, dict):
                    continue
                if not isinstance(site.get("folders"), list):
                    site["folders"] = []
                for folder_node in site.get("folders", []):
                    if not isinstance(folder_node, dict):
                        continue
                    if not isinstance(folder_node.get("files"), list):
                        folder_node["files"] = []
                    valid_files = []
                    for f in folder_node.get("files", []):
                        if not isinstance(f, dict):
                            continue
                        fn = str(f.get("filename") or "").strip()
                        if not fn:
                            print("[MonitoringServer] external_485 tree: filename 없는 항목 스킵", file=sys.stderr)
                            continue
                        f.setdefault("company", comp.get("company", "485"))
                        f.setdefault("folder", folder_node.get("folder", "REMOTE"))
                        f["source"] = "external_485"
                        valid_files.append(f)
                    folder_node["files"] = valid_files
        return tree

    # 케이스 B: {"files":[...]} 또는 [...] 형태를 표준 트리로 변환
    files = []
    if isinstance(payload, dict) and isinstance(payload.get("files"), list):
        files = payload.get("files", [])
    elif isinstance(payload, list):
        files = payload

    if not isinstance(files, list) or not files:
        return []

    grouped = {}
    for raw in files:
        if not isinstance(raw, dict):
            print("[MonitoringServer] external_485 files: dict 아님 스킵", file=sys.stderr)
            continue
        company = str(raw.get("company") or "485").strip() or "485"
        site = str(raw.get("site") or "외부485").strip() or "외부485"
        folder = str(raw.get("folder") or "REMOTE").strip() or "REMOTE"
        filename = str(raw.get("filename") or "").strip()
        if not filename:
            print("[MonitoringServer] external_485 files: filename 없음 스킵", file=sys.stderr)
            continue

        try:
            fill_interval = int(raw.get("fill_interval") or 0)
        except Exception:
            fill_interval = 0
            print(f"[MonitoringServer] external_485 files: fill_interval 파싱 실패 ({filename})", file=sys.stderr)
        try:
            gen_interval = int(raw.get("gen_interval") or 0)
        except Exception:
            gen_interval = 0
            print(f"[MonitoringServer] external_485 files: gen_interval 파싱 실패 ({filename})", file=sys.stderr)

        grouped.setdefault(company, {}).setdefault(site, {}).setdefault(folder, []).append({
            "filename": filename,
            "company": company,
            "folder": folder,
            "source": "external_485",
            "status": str(raw.get("status") or "외부"),
            "last_ts": str(raw.get("last_ts") or "-"),
            "elapsed": str(raw.get("elapsed") or "-"),
            "fill_interval": fill_interval,
            "gen_interval": gen_interval,
            "station_id": str(raw.get("station_id") or ""),
            "sensor_types": raw.get("sensor_types", []) if isinstance(raw.get("sensor_types"), list) else [],
        })

    tree = []
    for company, sites in grouped.items():
        company_node = {"company": company, "sites": [], "managers": []}
        for site, folders in sites.items():
            site_node = {
                "site": site,
                "folders": [],
                "site_address": "",
                "site_program": "",
                "check_interval": 0,
                "report_enabled": False,
                "report_cycle": "",
                "memo": "외부 485 목록",
                "assigned_manager_id": "",
                "assigned_qm_manager_id": "",
                "assigned_manager_name": "",
                "assigned_qm_manager_name": "",
                "station_groups": [],
                "stations": [],
            }
            for folder, file_items in folders.items():
                site_node["folders"].append({"folder": folder, "files": file_items})
            company_node["sites"].append(site_node)
        tree.append(company_node)
    return tree


# ─────────────────────────────────────────
#  API: 트리 전체
# ─────────────────────────────────────────

def _manager_name_by_id(managers, mgr_id):
    """담당자 id → 표시용 이름 (모니터링 트리용)."""
    if not mgr_id or not isinstance(managers, list):
        return ""
    for m in managers:
        if isinstance(m, dict) and m.get("id") == mgr_id:
            return (m.get("name") or "").strip()
    return ""


@app.route("/api/tree")
def api_tree():
    cfg  = _load_config()
    mgmt = _load_management()
    tree = []
    total = normal = delayed = unconverted = ghost = 0
    qm_list = mgmt.get("qm_managers", [])
    if not isinstance(qm_list, list):
        qm_list = []

    for company, company_val in cfg.items():
        if company.startswith("__") or not isinstance(company_val, dict):
            continue
        comp_mgmt = mgmt.get("companies", {}).get(company, {})
        comp_mgrs = comp_mgmt.get("managers", [])
        if not isinstance(comp_mgrs, list):
            comp_mgrs = []
        company_node = {
            "company": company,
            "sites": [],
            "managers": comp_mgrs,
        }

        for site, site_val in company_val.items():
            if site.startswith("__") or not isinstance(site_val, dict):
                continue
            site_mgmt   = comp_mgmt.get("sites", {}).get(site, _EMPTY_SITE_MGMT)
            assignments = site_mgmt.get("assignments", {})
            am_id = str(site_mgmt.get("assigned_manager_id") or "")
            aq_id = str(site_mgmt.get("assigned_qm_manager_id") or "")
            site_node = {
                "site": site,
                "folders": [],
                "site_address": str(site_mgmt.get("address") or "").strip(),
                "site_program": str(site_mgmt.get("program") or "").strip(),
                "check_interval": str(site_mgmt.get("check_interval") or ""),
                "report_enabled": bool(site_mgmt.get("report_enabled", False)),
                "report_cycle": str(site_mgmt.get("report_cycle") or ""),
                "memo": str(site_mgmt.get("memo") or ""),
                "assigned_manager_id": am_id,
                "assigned_qm_manager_id": aq_id,
                "assigned_manager_name": _manager_name_by_id(comp_mgrs, am_id),
                "assigned_qm_manager_name": _manager_name_by_id(qm_list, aq_id),
            }

            for folder, folder_val in site_val.items():
                if folder.startswith("__") or not isinstance(folder_val, dict):
                    continue
                folder_node = {"folder": folder, "files": []}
                is_ghost_folder = folder_val.get("__is_ghost__", False)

                for filename, file_cfg in folder_val.items():
                    if filename.startswith("__") or not isinstance(file_cfg, dict):
                        continue
                    is_ghost = is_ghost_folder or file_cfg.get("__is_ghost__", False)
                    out_path = os.path.join(CONVERT_ROOT, company, folder, filename)
                    st = _file_status(out_path, is_ghost)
                    total += 1
                    s = st["status"]
                    if s == "정상":    normal += 1
                    elif s == "지연":  delayed += 1
                    elif s == "Ghost": ghost += 1
                    else:              unconverted += 1

                    sensor_types = [
                        file_cfg.get(f"CH{i}", {}).get("sensor_type", "")
                        for i in range(8)
                        if file_cfg.get(f"CH{i}", {}).get("sensor_type", "")
                    ]
                    folder_node["files"].append({
                        "filename":      filename,
                        "company":       company,
                        "folder":        folder,
                        "source":        "local",
                        "status":        st["status"],
                        "last_ts":       st["last_ts"],
                        "elapsed":       st["elapsed"],
                        "fill_interval": file_cfg.get("__fill_interval__", 0),
                        "gen_interval":  file_cfg.get("__gen_interval__", 0),
                        "station_id":    assignments.get(f"{folder}/{filename}", ""),
                        "sensor_types":  sensor_types,
                    })

                if folder_node["files"]:
                    site_node["folders"].append(folder_node)
            # 로거 0건 현장도 대시보드/웹에 표시 (데스크톱에서 추가만 한 현장 포함)
            site_node["station_groups"] = _sorted_station_groups(site_mgmt)
            site_node["stations"] = _normalize_stations_for_api(site_mgmt.get("stations", []))
            company_node["sites"].append(site_node)
        if company_node["sites"]:
            tree.append(company_node)

    # 외부 485 목록 병합 (옵션)
    ext_tree = _fetch_external_485_tree(cfg)
    if ext_tree:
        for comp in ext_tree:
            if not isinstance(comp, dict):
                continue
            tree.append(comp)
            for site in comp.get("sites", []) or []:
                for folder in site.get("folders", []) or []:
                    for f in folder.get("files", []) or []:
                        total += 1
                        s = str(f.get("status") or "")
                        if s == "정상":
                            normal += 1
                        elif s == "지연":
                            delayed += 1
                        elif s == "Ghost":
                            ghost += 1
                        else:
                            unconverted += 1

    site_count = sum(len(c.get("sites", [])) for c in tree if isinstance(c, dict))
    company_count = len(tree)

    return jsonify({
        "tree": tree,
        "summary": {
            "total": total,
            "loggers": total,
            "sites": site_count,
            "companies": company_count,
            "normal": normal,
            "delayed": delayed,
            "unconverted": unconverted,
            "ghost": ghost,
        },
    })


# ─────────────────────────────────────────
#  API: 파일 상세 (최신값 + 차트 데이터 포함)
# ─────────────────────────────────────────

@app.route("/api/detail")
def api_detail():
    company  = request.args.get("company", "")
    folder   = request.args.get("folder", "")
    filename = request.args.get("filename", "")
    n_chart  = int(request.args.get("n", 48))

    cfg = _load_config()
    file_cfg = {}
    site_found = ""

    for site, site_val in cfg.get(company, {}).items():
        if site.startswith("__") or not isinstance(site_val, dict):
            continue
        if folder in site_val:
            file_cfg = site_val[folder].get(filename, {})
            site_found = site
            break

    is_ghost = file_cfg.get("__is_ghost__", False)
    out_path = os.path.join(CONVERT_ROOT, company, folder, filename)
    st = _file_status(out_path, is_ghost)
    channels = _parse_channels(file_cfg)

    # ── 차트/최신값: 캐시 우선, 없으면 CSV 직접 파싱 (초기 1회) ──
    try:
        from monitoring.data_cache import get_file_cache
        cached = get_file_cache(company, folder, filename)
    except Exception:
        cached = None

    if cached:
        latest_values = cached.get("latest_values", {f"CH{i}": None for i in range(8)})
        chart = cached.get("chart", {"labels": [], "values": {f"CH{i}": [] for i in range(8)}})
    else:
        # 캐시 없음 → CSV 직접 파싱 (변환 전 초기 상태)
        timestamps, ch_values = _read_csv_tail(out_path, n=n_chart)
        latest_values = {}
        for i in range(8):
            key = f"CH{i}"
            vals = [v for v in ch_values[key] if v is not None]
            latest_values[key] = vals[-1] if vals else None
        short_labels = []
        for ts in timestamps:
            try:
                dt = datetime.fromisoformat(ts)
                short_labels.append(dt.strftime("%m/%d %H:%M"))
            except Exception:
                short_labels.append(ts[-5:] if len(ts) >= 5 else ts)
        chart = {"labels": short_labels, "values": ch_values}

    return jsonify({
        "company": company,
        "site": site_found,
        "folder": folder,
        "filename": filename,
        "note": file_cfg.get("__note__", ""),
        "fill_interval": file_cfg.get("__fill_interval__", 0),
        "gen_interval": file_cfg.get("__gen_interval__", 0),
        "status": st["status"],
        "last_ts": st["last_ts"],
        "elapsed": st["elapsed"],
        "channels": channels,
        "latest_values": latest_values,
        "chart": chart,
        "from_cache": cached is not None,
    })


# ─────────────────────────────────────────
#  API: 채널 센서명 저장 (양방향)
# ─────────────────────────────────────────

@app.route("/api/channel", methods=["POST"])
def api_channel():
    data = request.get_json(force=True) or {}
    company     = data.get("company", "")
    folder      = data.get("folder", "")
    filename    = data.get("filename", "")
    ch          = data.get("ch", "")        # "CH0" ~ "CH7"
    label       = data.get("label")         # None이면 변경 안 함
    sensor_type = data.get("sensor_type")   # None이면 변경 안 함

    if ch not in [f"CH{i}" for i in range(8)]:
        return jsonify({"ok": False, "error": "잘못된 채널"}), 400

    try:
        cfg = _load_config()
        updated = False
        for site, site_val in cfg.get(company, {}).items():
            if site.startswith("__") or not isinstance(site_val, dict):
                continue
            if folder in site_val and filename in site_val[folder]:
                file_cfg = site_val[folder][filename]
                if ch not in file_cfg:
                    file_cfg[ch] = {}
                if label is not None:
                    file_cfg[ch]["label"] = label
                if sensor_type is not None:
                    file_cfg[ch]["sensor_type"] = sensor_type
                updated = True
                break
        if not updated:
            return jsonify({"ok": False, "error": "파일을 찾을 수 없습니다"}), 404
        _save_config(cfg)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 비고 저장 (양방향)
# ─────────────────────────────────────────

@app.route("/api/note", methods=["POST"])
def api_note():
    data = request.get_json(force=True) or {}
    company  = data.get("company", "")
    folder   = data.get("folder", "")
    filename = data.get("filename", "")
    note     = data.get("note", "")

    try:
        cfg = _load_config()
        updated = False
        for site, site_val in cfg.get(company, {}).items():
            if site.startswith("__") or not isinstance(site_val, dict):
                continue
            if folder in site_val and filename in site_val[folder]:
                site_val[folder][filename]["__note__"] = note
                updated = True
                break
        if not updated:
            return jsonify({"ok": False, "error": "파일을 찾을 수 없습니다"}), 404
        _save_config(cfg)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 업체 목록 + 담당자 조회
# ─────────────────────────────────────────

@app.route("/api/companies")
def api_companies():
    cfg  = _load_config()
    mgmt = _load_management()
    qm_managers = mgmt.get("qm_managers", [])
    companies = []

    for company, company_val in cfg.items():
        if company.startswith("__") or not isinstance(company_val, dict):
            continue
        comp_mgmt = mgmt.get("companies", {}).get(company, {})
        managers  = comp_mgmt.get("managers", [])
        sites = []

        for site, site_val in company_val.items():
            if site.startswith("__") or not isinstance(site_val, dict):
                continue
            site_mgmt   = comp_mgmt.get("sites", {}).get(site, _EMPTY_SITE_MGMT)
            stations    = site_mgmt.get("stations", [])
            assignments = site_mgmt.get("assignments", {})

            # 개소별 로거 수 계산
            station_logger_count = {st["id"]: 0 for st in stations if isinstance(st, dict)}
            total_loggers = 0
            for folder, folder_val in site_val.items():
                if folder.startswith("__") or not isinstance(folder_val, dict):
                    continue
                for filename, file_cfg in folder_val.items():
                    if filename.startswith("__") or not isinstance(file_cfg, dict):
                        continue
                    total_loggers += 1
                    sid = assignments.get(f"{folder}/{filename}", "")
                    if sid in station_logger_count:
                        station_logger_count[sid] += 1

            norm_st = _normalize_stations_for_api(stations)
            enriched_stations = [
                {**st, "logger_count": station_logger_count.get(st["id"], 0)}
                for st in norm_st
            ]
            sites.append({
                "site":                    site,
                "site_address":            str(site_mgmt.get("address") or "").strip(),
                "site_program":            str(site_mgmt.get("program") or "").strip(),
                "check_interval":          str(site_mgmt.get("check_interval") or ""),
                "report_enabled":          site_mgmt.get("report_enabled", False),
                "report_cycle":            site_mgmt.get("report_cycle", ""),
                "memo":                    site_mgmt.get("memo", ""),
                "station_groups":          _sorted_station_groups(site_mgmt),
                "stations":                enriched_stations,
                "total_loggers":           total_loggers,
                "assigned_manager_id":     site_mgmt.get("assigned_manager_id", ""),
                "assigned_qm_manager_id":  site_mgmt.get("assigned_qm_manager_id", ""),
            })
        companies.append({"company": company, "managers": managers, "sites": sites})
    return jsonify({
        "companies": companies,
        "qm_managers": qm_managers,
        "program_types": _get_program_types_list(mgmt),
    })


@app.route("/api/settings/program_types", methods=["GET"])
def api_program_types_get():
    mgmt = _load_management()
    return jsonify({"ok": True, "programs": _get_program_types_list(mgmt)})


@app.route("/api/settings/program_types", methods=["POST"])
def api_program_types_post():
    """management.json program_types 갱신 (웹 관리자 설정)."""
    data = request.get_json(force=True) or {}
    programs = data.get("programs")
    if not isinstance(programs, list):
        return jsonify({"ok": False, "error": "programs 배열이 필요합니다"}), 400
    seen = set()
    cleaned = []
    for x in programs:
        s = str(x).strip()
        if s and s not in seen:
            seen.add(s)
            cleaned.append(s)
    if not cleaned:
        return jsonify({"ok": False, "error": "최소 하나의 프로그램명을 입력하세요"}), 400
    try:
        with _edit_management() as mgmt:
            mgmt["program_types"] = cleaned
        return jsonify({"ok": True, "programs": cleaned})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  개소 헬퍼: 현장 찾기
# ─────────────────────────────────────────

def _find_site(cfg, company, site):
    """config.json cfg[company][site] 반환. 없으면 None."""
    company_val = cfg.get(company)
    if not isinstance(company_val, dict):
        return None
    site_val = company_val.get(site)
    return site_val if isinstance(site_val, dict) else None


def _sorted_station_groups(site_mgmt: dict) -> list:
    """대분류(station_groups) 정렬 목록. management.json의 현장 dict를 받는다."""
    raw = site_mgmt.get("station_groups", [])
    if not isinstance(raw, list):
        return []
    out = []
    for g in raw:
        if isinstance(g, dict) and g.get("id"):
            out.append({
                "id": g["id"],
                "name": str(g.get("name") or ""),
                "order": int(g.get("order") or 0),
            })
    return sorted(out, key=lambda x: x["order"])


def _normalize_stations_for_api(stations) -> list:
    """소분류 목록에 group_id 문자열 보장. management.json의 stations 리스트를 받는다."""
    if not isinstance(stations, list):
        return []
    out = []
    for st in stations:
        if not isinstance(st, dict) or not st.get("id"):
            continue
        d = dict(st)
        d["group_id"] = str(d.get("group_id") or "")
        out.append(d)
    return out


# ─────────────────────────────────────────
#  API: 대분류(상위 카테고리) 추가
# ─────────────────────────────────────────

@app.route("/api/station/group", methods=["POST"])
def api_station_group_add():
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    site    = data.get("site", "")
    name    = (data.get("name") or "").strip()

    if not company or not site or not name:
        return jsonify({"ok": False, "error": "company/site/name 필수"}), 400

    try:
        rec = {}
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            groups = site_mgmt.setdefault("station_groups", [])
            if not isinstance(groups, list):
                groups = site_mgmt["station_groups"] = []
            new_id = "grp_" + str(int(time.time() * 1000))
            order = max((g.get("order", 0) for g in groups if isinstance(g, dict)), default=-1) + 1
            rec = {"id": new_id, "name": name, "order": order}
            groups.append(rec)
        return jsonify({"ok": True, "group": rec})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/station/group", methods=["DELETE"])
def api_station_group_delete():
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    site    = data.get("site", "")
    grp_id  = data.get("id", "")

    try:
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            groups = site_mgmt.get("station_groups", [])
            if isinstance(groups, list):
                site_mgmt["station_groups"] = [g for g in groups if isinstance(g, dict) and g.get("id") != grp_id]
            for st in site_mgmt.get("stations", []) or []:
                if isinstance(st, dict) and st.get("group_id") == grp_id:
                    st["group_id"] = ""
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/station/group/rename", methods=["POST"])
def api_station_group_rename():
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    site    = data.get("site", "")
    grp_id  = data.get("id", "")
    name    = (data.get("name") or "").strip()

    if not name:
        return jsonify({"ok": False, "error": "이름을 입력하세요"}), 400

    try:
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            for g in site_mgmt.get("station_groups", []) or []:
                if isinstance(g, dict) and g.get("id") == grp_id:
                    g["name"] = name
                    break
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/station/group/reorder", methods=["POST"])
def api_station_group_reorder():
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    site    = data.get("site", "")
    ids     = data.get("ids", [])

    try:
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            id_to_order = {gid: idx for idx, gid in enumerate(ids)}
            for g in site_mgmt.get("station_groups", []) or []:
                if isinstance(g, dict) and g.get("id") in id_to_order:
                    g["order"] = id_to_order[g["id"]]
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 개소 추가
# ─────────────────────────────────────────

@app.route("/api/station", methods=["POST"])
def api_station_add():
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    site    = data.get("site", "")
    name    = (data.get("name") or "").strip()
    group_id = str(data.get("group_id") or "")

    if not company or not site or not name:
        return jsonify({"ok": False, "error": "company/site/name 필수"}), 400

    try:
        rec = {}
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            stations = site_mgmt.setdefault("stations", [])
            new_id = "st_" + str(int(time.time() * 1000))
            order  = max((s.get("order", 0) for s in stations), default=-1) + 1
            rec = {"id": new_id, "name": name, "order": order, "group_id": group_id}
            stations.append(rec)
        return jsonify({"ok": True, "station": {**rec, "logger_count": 0}})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 개소 삭제
# ─────────────────────────────────────────

@app.route("/api/station", methods=["DELETE"])
def api_station_delete():
    data       = request.get_json(force=True) or {}
    company    = data.get("company", "")
    site       = data.get("site", "")
    station_id = data.get("id", "")

    try:
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            site_mgmt["stations"] = [
                s for s in site_mgmt.get("stations", []) if s.get("id") != station_id
            ]
            # 배정 해제: assignments에서 해당 station_id 제거
            assignments = site_mgmt.get("assignments", {})
            for key, sid in list(assignments.items()):
                if sid == station_id:
                    assignments[key] = ""
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 개소 이름 변경
# ─────────────────────────────────────────

# ─────────────────────────────────────────
#  API: 개소 순서 변경
# ─────────────────────────────────────────

@app.route("/api/station/reorder", methods=["POST"])
def api_station_reorder():
    """body: {company, site, ids: ["st_1", "st_2", ...]}  → order를 인덱스 순으로 갱신"""
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    site    = data.get("site", "")
    ids     = data.get("ids", [])

    try:
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            id_to_order = {sid: idx for idx, sid in enumerate(ids)}
            for st in site_mgmt.get("stations", []):
                if st.get("id") in id_to_order:
                    st["order"] = id_to_order[st["id"]]
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 개소 이름 변경
# ─────────────────────────────────────────

@app.route("/api/station/update_group", methods=["POST"])
def api_station_update_group():
    """미분류 소분류를 대분류(group)에 배정하거나 해제한다. group_id="" → 미분류."""
    data       = request.get_json(force=True) or {}
    company    = data.get("company", "")
    site       = data.get("site", "")
    station_id = data.get("id", "")
    group_id   = str(data.get("group_id") or "")

    if not company or not site or not station_id:
        return jsonify({"ok": False, "error": "company/site/id 필수"}), 400

    try:
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            for st in site_mgmt.get("stations", []):
                if st.get("id") == station_id:
                    st["group_id"] = group_id
                    break
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/station/rename", methods=["POST"])
def api_station_rename():
    data       = request.get_json(force=True) or {}
    company    = data.get("company", "")
    site       = data.get("site", "")
    station_id = data.get("id", "")
    new_name   = (data.get("name") or "").strip()

    if not new_name:
        return jsonify({"ok": False, "error": "이름을 입력하세요"}), 400

    try:
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            for st in site_mgmt.get("stations", []):
                if st.get("id") == station_id:
                    st["name"] = new_name
                    break
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 로거 개소 배정 (복수 일괄)
# ─────────────────────────────────────────

@app.route("/api/station/save_all", methods=["POST"])
def api_station_save_all():
    """카테고리 전체 일괄 저장.
    station_groups, stations, assignments 를 management.json에 덮어씌운다.
    웹 UI의 저장 버튼에서 호출.
    """
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    site    = data.get("site", "")

    if not company or not site:
        return jsonify({"ok": False, "error": "company/site 필수"}), 400

    cfg = _load_config()
    if _find_site(cfg, company, site) is None:
        return jsonify({"ok": False, "error": "config.json에 없는 현장입니다"}), 404

    try:
        with _edit_management() as mgmt:
            # management.json에 현장 노드가 아직 없으면 생성 (신규 현장·구버전 데이터)
            site_mgmt = _ensure_site_mgmt(mgmt, company, site)
            site_mgmt["station_groups"] = data.get("station_groups", [])
            site_mgmt["stations"]       = data.get("stations", [])
            site_mgmt["assignments"]    = data.get("assignments", {})
        # 데스크탑 앱 트리 새로고침 신호
        if _on_category_saved:
            try:
                _on_category_saved()
            except Exception:
                pass
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/station/assign", methods=["POST"])
def api_station_assign():
    """
    body: {
      company, site, station_id,
      loggers: [{"folder": "...", "filename": "..."}, ...]
    }
    station_id == "" → 미배정
    loggers 목록에 있는 파일은 station_id로 설정,
    같은 현장의 나머지 파일 중 현재 station_id인 것은 미배정.
    """
    data       = request.get_json(force=True) or {}
    company    = data.get("company", "")
    site       = data.get("site", "")
    station_id = data.get("station_id", "")
    loggers    = data.get("loggers", [])   # [{"folder":…,"filename":…}]

    try:
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")

            assignments   = site_mgmt.setdefault("assignments", {})
            assigned_keys = {(lg["folder"], lg["filename"]) for lg in loggers}

            # 이 개소에 현재 배정된 것 중 loggers 목록에 없는 건 미배정으로
            for key, sid in list(assignments.items()):
                if sid == station_id and station_id:
                    folder_part, fname_part = key.split("/", 1)
                    if (folder_part, fname_part) not in assigned_keys:
                        assignments[key] = ""

            # loggers 목록을 station_id로 배정
            for lg in loggers:
                assignments[f"{lg['folder']}/{lg['filename']}"] = station_id

        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 로거 단일 개소 배정 / 해제
# ─────────────────────────────────────────

@app.route("/api/logger/station", methods=["POST"])
def api_logger_station():
    """단일 로거의 __station_id__ 설정. station_id="" → 미배정."""
    data       = request.get_json(force=True) or {}
    company    = data.get("company", "")
    site       = data.get("site", "")
    folder     = data.get("folder", "")
    filename   = data.get("filename", "")
    station_id = data.get("station_id", "")

    try:
        with _edit_management() as mgmt:
            site_mgmt = _find_site_mgmt(mgmt, company, site)
            if site_mgmt is None:
                raise LookupError("현장을 찾을 수 없습니다")
            assignments = site_mgmt.setdefault("assignments", {})
            assignments[f"{folder}/{filename}"] = station_id
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 현장 내 전체 로거 목록 (개소 배정용)
# ─────────────────────────────────────────

@app.route("/api/site/loggers")
def api_site_loggers():
    company = request.args.get("company", "")
    site    = request.args.get("site", "")

    cfg      = _load_config()
    mgmt     = _load_management()
    site_val = _find_site(cfg, company, site)
    if site_val is None:
        return jsonify({"loggers": []})

    site_mgmt   = _find_site_mgmt(mgmt, company, site) or {}
    assignments = site_mgmt.get("assignments", {})

    loggers = []
    for folder, folder_val in site_val.items():
        if folder.startswith("__") or not isinstance(folder_val, dict):
            continue
        for filename, file_cfg in folder_val.items():
            if filename.startswith("__") or not isinstance(file_cfg, dict):
                continue
            loggers.append({
                "folder":     folder,
                "filename":   filename,
                "station_id": assignments.get(f"{folder}/{filename}", ""),
                "label":      file_cfg.get("CH0", {}).get("label", "") or filename,
            })
    loggers.sort(key=lambda x: (x["folder"], x["filename"]))
    return jsonify({"loggers": loggers})


# ─────────────────────────────────────────
#  API: 담당자 추가 / 수정
# ─────────────────────────────────────────

@app.route("/api/company/manager", methods=["POST"])
def api_manager_save():
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    name    = (data.get("name") or "").strip()
    title   = (data.get("title") or "").strip()
    phone   = (data.get("phone") or "").strip()
    mgr_id  = (data.get("id") or "").strip()

    if not company or not name:
        return jsonify({"ok": False, "error": "업체명과 이름은 필수입니다"}), 400

    try:
        result_managers = []
        with _edit_management() as mgmt:
            cfg = _load_config()
            if company not in cfg or not isinstance(cfg.get(company), dict):
                raise LookupError("업체를 찾을 수 없습니다")
            comp = mgmt.setdefault("companies", {}).setdefault(
                company, {"managers": [], "sites": {}}
            )
            managers = comp.setdefault("managers", [])
            if mgr_id:
                for i, m in enumerate(managers):
                    if m.get("id") == mgr_id:
                        managers[i] = {"id": mgr_id, "name": name, "title": title, "phone": phone}
                        break
                else:
                    managers.append({"id": mgr_id, "name": name, "title": title, "phone": phone})
            else:
                new_id = str(int(time.time() * 1000))
                managers.append({"id": new_id, "name": name, "title": title, "phone": phone})
            result_managers = managers
        return jsonify({"ok": True, "managers": result_managers})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 업체 신규 추가 (빈 업체 → 이후 현장 추가)
# ─────────────────────────────────────────

@app.route("/api/company/create", methods=["POST"])
def api_company_create():
    data    = request.get_json(force=True) or {}
    company = (data.get("company") or "").strip()

    if not company:
        return jsonify({"ok": False, "error": "업체명은 필수입니다"}), 400
    if company.startswith("__"):
        return jsonify({"ok": False, "error": "업체명은 __ 로 시작할 수 없습니다"}), 400

    try:
        with _edit_config() as cfg:
            if company in cfg and isinstance(cfg.get(company), dict):
                raise ValueError("이미 같은 이름의 업체가 있습니다")
            cfg[company] = {}

        with _edit_management() as mgmt:
            mgmt.setdefault("companies", {}).setdefault(
                company, {"managers": [], "sites": {}}
            )

        if _on_category_saved:
            try:
                _on_category_saved()
            except Exception:
                pass
        return jsonify({"ok": True, "company": company})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 409
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 현장 신규 추가 (웹 업체 관리)
# ─────────────────────────────────────────

@app.route("/api/site/create", methods=["POST"])
def api_site_create():
    data    = request.get_json(force=True) or {}
    company = (data.get("company") or "").strip()
    site    = (data.get("site") or "").strip()
    s       = data.get("settings") or {}

    if not company or not site:
        return jsonify({"ok": False, "error": "업체명과 현장명은 필수입니다"}), 400
    if site.startswith("__"):
        return jsonify({"ok": False, "error": "현장명은 __ 로 시작할 수 없습니다"}), 400

    try:
        site_mgmt_out = {}
        # config.json에 빈 현장 노드 생성 (폴더/파일 없는 빈 현장)
        with _edit_config() as cfg:
            company_val = cfg.get(company)
            if not isinstance(company_val, dict):
                raise LookupError("업체를 찾을 수 없습니다")
            if site in company_val:
                raise ValueError("같은 이름의 현장이 이미 있습니다")
            company_val[site] = {"__note__": ""}

        # management.json에 현장 관리 데이터 저장
        with _edit_management() as mgmt:
            site_mgmt = _ensure_site_mgmt(mgmt, company, site)
            site_mgmt["address"]                 = str(s.get("site_address") or "").strip()
            site_mgmt["program"]                 = str(s.get("site_program") or "").strip()
            site_mgmt["check_interval"]          = str(s.get("check_interval") or "").strip()
            site_mgmt["report_enabled"]          = bool(s.get("report_enabled", False))
            site_mgmt["report_cycle"]            = str(s.get("report_cycle") or "").strip()
            site_mgmt["memo"]                    = str(s.get("memo") or "").strip()
            site_mgmt["assigned_manager_id"]     = str(s.get("assigned_manager_id") or "").strip()
            site_mgmt["assigned_qm_manager_id"]  = str(s.get("assigned_qm_manager_id") or "").strip()
            site_mgmt_out = site_mgmt

        out_site = {
            "site":                    site,
            "site_address":            site_mgmt_out.get("address", ""),
            "site_program":            site_mgmt_out.get("program", ""),
            "check_interval":          str(site_mgmt_out.get("check_interval") or ""),
            "report_enabled":          site_mgmt_out.get("report_enabled", False),
            "report_cycle":            site_mgmt_out.get("report_cycle", ""),
            "memo":                    site_mgmt_out.get("memo", ""),
            "station_groups":          _sorted_station_groups(site_mgmt_out),
            "stations":                [],
            "total_loggers":           0,
            "assigned_manager_id":     site_mgmt_out.get("assigned_manager_id", ""),
            "assigned_qm_manager_id":  site_mgmt_out.get("assigned_qm_manager_id", ""),
        }
        if _on_category_saved:
            try:
                _on_category_saved()
            except Exception:
                pass
        return jsonify({"ok": True, "site": out_site})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 409
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/site/delete", methods=["POST"])
def api_site_delete():
    """config.json·management.json에서 현장 노드 제거. 복구 불가."""
    data    = request.get_json(force=True) or {}
    company = (data.get("company") or "").strip()
    site    = (data.get("site") or "").strip()

    if not company or not site:
        return jsonify({"ok": False, "error": "업체명과 현장명은 필수입니다"}), 400
    if site.startswith("__"):
        return jsonify({"ok": False, "error": "유효하지 않은 현장명입니다"}), 400

    try:
        with _edit_config() as cfg:
            company_val = cfg.get(company)
            if not isinstance(company_val, dict):
                raise LookupError("업체를 찾을 수 없습니다")
            if site not in company_val:
                raise LookupError("현장을 찾을 수 없습니다")
            del company_val[site]

        with _edit_management() as mgmt:
            comp = mgmt.get("companies", {}).get(company)
            if comp and isinstance(comp.get("sites"), dict) and site in comp["sites"]:
                del comp["sites"][site]

        if _on_category_saved:
            try:
                _on_category_saved()
            except Exception:
                pass
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 현장 설정 저장
# ─────────────────────────────────────────

@app.route("/api/site/settings", methods=["POST"])
def api_site_settings():
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    site    = data.get("site", "")
    s       = data.get("settings", {})

    if not company or not site:
        return jsonify({"ok": False, "error": "업체명과 현장명은 필수입니다"}), 400

    try:
        with _edit_management() as mgmt:
            site_mgmt = _ensure_site_mgmt(mgmt, company, site)
            if "site_address" in s:
                site_mgmt["address"] = str(s["site_address"] or "").strip()
            if "site_program" in s:
                site_mgmt["program"] = str(s["site_program"] or "").strip()
            if "check_interval" in s:
                site_mgmt["check_interval"] = str(s["check_interval"] or "").strip()
            if "report_enabled" in s:
                site_mgmt["report_enabled"] = bool(s["report_enabled"])
            if "report_cycle" in s:
                site_mgmt["report_cycle"] = str(s["report_cycle"]).strip()
            if "memo" in s:
                site_mgmt["memo"] = str(s["memo"]).strip()
            if "assigned_manager_id" in s:
                site_mgmt["assigned_manager_id"] = str(s["assigned_manager_id"]).strip()
            if "assigned_qm_manager_id" in s:
                site_mgmt["assigned_qm_manager_id"] = str(s["assigned_qm_manager_id"]).strip()
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: QM 관리자 추가
# ─────────────────────────────────────────

@app.route("/api/qm/manager", methods=["POST"])
def api_qm_manager_save():
    data   = request.get_json(force=True) or {}
    name   = (data.get("name") or "").strip()
    title  = (data.get("title") or "").strip()
    phone  = (data.get("phone") or "").strip()
    mgr_id = (data.get("id") or "").strip()

    if not name:
        return jsonify({"ok": False, "error": "이름은 필수입니다"}), 400

    try:
        result_managers = []
        with _edit_management() as mgmt:
            managers = mgmt.setdefault("qm_managers", [])
            if mgr_id:
                for i, m in enumerate(managers):
                    if m.get("id") == mgr_id:
                        managers[i] = {"id": mgr_id, "name": name, "title": title, "phone": phone}
                        break
                else:
                    managers.append({"id": mgr_id, "name": name, "title": title, "phone": phone})
            else:
                new_id = "qm_" + str(int(time.time() * 1000))
                managers.append({"id": new_id, "name": name, "title": title, "phone": phone})
            result_managers = managers
        return jsonify({"ok": True, "qm_managers": result_managers})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: QM 관리자 삭제
# ─────────────────────────────────────────

@app.route("/api/qm/manager", methods=["DELETE"])
def api_qm_manager_delete():
    data   = request.get_json(force=True) or {}
    mgr_id = data.get("id", "")

    try:
        result_managers = []
        with _edit_management() as mgmt:
            mgmt["qm_managers"] = [
                m for m in mgmt.get("qm_managers", [])
                if m.get("id") != mgr_id
            ]
            # 해당 QM 관리자가 배정된 현장 초기화
            for comp in mgmt.get("companies", {}).values():
                for site_mgmt in comp.get("sites", {}).values():
                    if site_mgmt.get("assigned_qm_manager_id") == mgr_id:
                        site_mgmt["assigned_qm_manager_id"] = ""
            result_managers = mgmt.get("qm_managers", [])
        return jsonify({"ok": True, "qm_managers": result_managers})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  API: 담당자 삭제
# ─────────────────────────────────────────

@app.route("/api/company/manager", methods=["DELETE"])
def api_manager_delete():
    data    = request.get_json(force=True) or {}
    company = data.get("company", "")
    mgr_id  = data.get("id", "")

    try:
        with _edit_management() as mgmt:
            cfg = _load_config()
            if company not in cfg or not isinstance(cfg.get(company), dict):
                raise LookupError("업체를 찾을 수 없습니다")
            comp = mgmt.get("companies", {}).get(company, {})
            comp["managers"] = [
                m for m in comp.get("managers", [])
                if m.get("id") != mgr_id
            ]
            # 해당 담당자가 배정된 현장 초기화
            for site_mgmt in comp.get("sites", {}).values():
                if site_mgmt.get("assigned_manager_id") == mgr_id:
                    site_mgmt["assigned_manager_id"] = ""
        return jsonify({"ok": True})
    except LookupError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────
#  QM 원격 — 로컬 SQLite (qm_remote_state.sqlite3)
#  qm_remote.json: servers[], request_timeout_sec, session_admin_password, desktop_api_token
# ─────────────────────────────────────────

QM_REMOTE_DEFAULT_SESSION_ADMIN_PW = "1524"


def _qm_effective_session_admin_pw(raw) -> str:
    """비밀번호 미설정·빈 문자열이면 공통 기본값."""
    s = str(raw or "").strip()[:120]
    return s if s else QM_REMOTE_DEFAULT_SESSION_ADMIN_PW


def _default_qm_remote_config() -> dict:
    return {
        "servers": [
            "큐엠메인서버1",
            "큐엠메인서버2",
            "큐엠메인서버3",
            "큐엠메인서버5",
        ],
        "request_timeout_sec": 5.0,
        "session_admin_password": QM_REMOTE_DEFAULT_SESSION_ADMIN_PW,
        "desktop_api_token": "",
    }


def _load_qm_remote_config() -> dict:
    dflt = _default_qm_remote_config()
    if not QM_REMOTE_PATH.exists():
        try:
            with open(QM_REMOTE_PATH, "w", encoding="utf-8") as f:
                json.dump(dflt, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        return dflt.copy()
    try:
        with open(QM_REMOTE_PATH, encoding="utf-8") as f:
            raw = json.load(f)
        if not isinstance(raw, dict):
            return dflt.copy()
        try:
            rts = float(raw.get("request_timeout_sec", 5))
            rts = max(1.0, min(30.0, rts))
        except Exception:
            rts = 5.0
        servers: list[str] = []
        rs = raw.get("servers")
        if isinstance(rs, list):
            servers = [str(x).strip() for x in rs if str(x).strip()]
        if not servers:
            servers = list(dflt["servers"])
        return {
            "request_timeout_sec": rts,
            "servers": servers,
            "session_admin_password": _qm_effective_session_admin_pw(raw.get("session_admin_password")),
            "desktop_api_token": str(raw.get("desktop_api_token") or "")[:200],
        }
    except Exception:
        return dflt.copy()


def _save_qm_remote_config(data: dict) -> None:
    dflt = _default_qm_remote_config()
    try:
        rts = float(data.get("request_timeout_sec", 5))
        rts = max(1.0, min(30.0, rts))
    except Exception:
        rts = 5.0
    servers: list[str] = []
    if isinstance(data.get("servers"), list):
        servers = [str(x).strip() for x in data["servers"] if str(x).strip()]
    if not servers:
        servers = list(dflt["servers"])
    sap = _qm_effective_session_admin_pw(data.get("session_admin_password"))
    payload = {
        "request_timeout_sec": rts,
        "servers": servers,
        "session_admin_password": sap,
        "desktop_api_token": str(data.get("desktop_api_token") or "")[:200],
    }
    with open(QM_REMOTE_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _qm_init_local_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS qm_server_state (
          server_name TEXT PRIMARY KEY,
          status TEXT NOT NULL DEFAULT 'OFF',
          user TEXT NOT NULL DEFAULT '',
          timestamp TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS qm_server_note (
          server_name TEXT PRIMARY KEY,
          note TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS qm_pc_display (
          pc_id TEXT PRIMARY KEY,
          display_name TEXT NOT NULL DEFAULT ''
        );
        """
    )


def _qm_local_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(QM_LOCAL_DB_PATH), timeout=30.0)
    conn.row_factory = sqlite3.Row
    _qm_init_local_db(conn)
    conn.commit()
    return conn


def _qm_local_get_state(server_name: str) -> dict:
    sn = (server_name or "").strip()
    if not sn:
        return {"status": "OFF", "user": "", "timestamp": ""}
    with _QM_DB_LOCK:
        c = _qm_local_conn()
        try:
            row = c.execute(
                "SELECT status, user, timestamp FROM qm_server_state WHERE server_name = ?",
                (sn,),
            ).fetchone()
            if not row:
                return {"status": "OFF", "user": "", "timestamp": ""}
            return {
                "status": str(row["status"] or "OFF").upper(),
                "user": str(row["user"] or ""),
                "timestamp": str(row["timestamp"] or ""),
            }
        finally:
            c.close()


def _qm_local_set_state(server_name: str, state: dict) -> None:
    sn = (server_name or "").strip()
    if not sn:
        return
    st = _qm_normalize_state(state)
    with _QM_DB_LOCK:
        c = _qm_local_conn()
        try:
            c.execute(
                """INSERT INTO qm_server_state(server_name,status,user,timestamp)
                   VALUES(?,?,?,?)
                   ON CONFLICT(server_name) DO UPDATE SET
                   status=excluded.status,user=excluded.user,timestamp=excluded.timestamp""",
                (sn, st["status"], st["user"], st["timestamp"]),
            )
            c.commit()
        finally:
            c.close()


def _qm_local_all_states(server_names: list[str]) -> dict:
    out: dict[str, dict] = {}
    with _QM_DB_LOCK:
        c = _qm_local_conn()
        try:
            for name in server_names:
                n = (name or "").strip()
                if not n:
                    continue
                row = c.execute(
                    "SELECT status, user, timestamp FROM qm_server_state WHERE server_name = ?",
                    (n,),
                ).fetchone()
                if not row:
                    out[n] = {"status": "OFF", "user": "", "timestamp": ""}
                else:
                    out[n] = {
                        "status": str(row["status"] or "OFF").upper(),
                        "user": str(row["user"] or ""),
                        "timestamp": str(row["timestamp"] or ""),
                    }
            return out
        finally:
            c.close()


def _qm_local_get_note(server_name: str) -> str:
    sn = (server_name or "").strip()
    if not sn:
        return ""
    with _QM_DB_LOCK:
        c = _qm_local_conn()
        try:
            row = c.execute(
                "SELECT note FROM qm_server_note WHERE server_name = ?", (sn,)
            ).fetchone()
            return str(row["note"] or "") if row else ""
        finally:
            c.close()


def _qm_local_all_notes(server_names: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    with _QM_DB_LOCK:
        c = _qm_local_conn()
        try:
            for n in server_names:
                sn = (n or "").strip()
                if not sn:
                    continue
                row = c.execute(
                    "SELECT note FROM qm_server_note WHERE server_name = ?", (sn,)
                ).fetchone()
                out[sn] = str(row["note"] or "") if row else ""
            return out
        finally:
            c.close()


def _qm_local_set_note(server_name: str, note: str) -> None:
    sn = (server_name or "").strip()
    if not sn:
        return
    note = str(note or "")[:2000]
    with _QM_DB_LOCK:
        c = _qm_local_conn()
        try:
            c.execute(
                """INSERT INTO qm_server_note(server_name,note) VALUES(?,?)
                   ON CONFLICT(server_name) DO UPDATE SET note=excluded.note""",
                (sn, note),
            )
            c.commit()
        finally:
            c.close()


def _qm_local_get_display_name(pc_id: str) -> str:
    pid = (pc_id or "").strip() or "UNKNOWN_PC"
    with _QM_DB_LOCK:
        c = _qm_local_conn()
        try:
            row = c.execute(
                "SELECT display_name FROM qm_pc_display WHERE pc_id = ?", (pid,)
            ).fetchone()
            return str(row["display_name"] or "").strip()[:200] if row else ""
        finally:
            c.close()


def _qm_local_set_display_name(pc_id: str, name: str) -> None:
    pid = (pc_id or "").strip() or "UNKNOWN_PC"
    name = str(name or "").strip()[:200]
    with _QM_DB_LOCK:
        c = _qm_local_conn()
        try:
            c.execute(
                """INSERT INTO qm_pc_display(pc_id,display_name) VALUES(?,?)
                   ON CONFLICT(pc_id) DO UPDATE SET display_name=excluded.display_name""",
                (pid, name),
            )
            c.commit()
        finally:
            c.close()


def _qm_session_effective_user(cfg: dict, pc_id: str, timeout: float) -> str:
    _ = cfg, timeout
    d = _qm_local_get_display_name(pc_id)
    return d if d else pc_id


def _qm_find_other_room_occupied_by_user(servers_ordered: list, me: str, exclude_server: str):
    """동일 사용자(me)가 다른 방(메인서버)에서 이미 ON이면 그 방 이름 반환, 없으면 None."""
    if not me:
        return None
    ex = (exclude_server or "").strip()
    for raw in servers_ordered or []:
        sn = str(raw or "").strip()
        if not sn or sn == ex:
            continue
        ost = _qm_normalize_state(_qm_local_get_state(sn))
        if ost["status"] == "ON" and ost["user"] == me:
            return sn
    return None


def _qm_pc_id() -> str:
    v = (os.environ.get("COMPUTERNAME") or os.environ.get("HOSTNAME") or "").strip()
    if v:
        return v
    try:
        import socket

        return (socket.gethostname() or "UNKNOWN_PC").split(".")[0]
    except Exception:
        return "UNKNOWN_PC"


def _qm_client_pc_id() -> str:
    """브라우저·데스크톱이 보낸 PC 식별자(윈도우 컴퓨터 이름 권장). 없으면 서버 호스트명."""
    h = (request.headers.get("X-QM-Client-PC-ID") or "").strip()
    if h:
        return h[:120]
    if request.method == "GET":
        q = (request.args.get("pc_id") or "").strip()
        if q:
            return q[:120]
    data = request.get_json(force=True, silent=True)
    if isinstance(data, dict):
        p = str(data.get("pc_id") or "").strip()
        if p:
            return p[:120]
    return _qm_pc_id()


def _qm_normalize_state(raw) -> dict:
    if not isinstance(raw, dict):
        return {"status": "OFF", "user": "", "timestamp": ""}
    st = str(raw.get("status") or "OFF").upper()
    if st not in ("ON", "OFF"):
        st = "OFF"
    ts_raw = raw.get("timestamp")
    ts = _qm_coerce_timestamp_string(ts_raw)
    return {
        "status": st,
        "user": str(raw.get("user") or "").strip(),
        "timestamp": ts,
    }


def _qm_coerce_timestamp_string(ts_raw) -> str:
    """타임스탬프 원시값을 '분석용 문자열'로 통일."""
    if ts_raw is None:
        return ""
    if isinstance(ts_raw, (int, float)):
        try:
            v = float(ts_raw)
            if v > 1e12:
                v = v / 1000.0
            return datetime.fromtimestamp(v).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(ts_raw).strip()
    if isinstance(ts_raw, str):
        return ts_raw.strip()
    return str(ts_raw).strip()


def _qm_parse_started_at(ts: str) -> tuple[datetime | None, str]:
    """
    시작 시각 문자열 → (로컬 naive datetime | None, 화면에 보여 줄 문구).
    파싱 실패 시 (None, 원문 또는 안내).
    """
    s = (ts or "").strip()
    if not s:
        return None, ""
    fmts = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    )
    for fmt in fmts:
        try:
            dt = datetime.strptime(s, fmt)
            return dt, dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    iso = s.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(iso)
        if dt.tzinfo is not None:
            dt = dt.astimezone().replace(tzinfo=None)
        return dt, dt.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        pass
    return None, s


def _qm_format_elapsed(start: datetime, now: datetime | None = None) -> str:
    """경과 시간 — 데스크톱은 분만 썼으나 24시간 넘으면 일/시간도 표시."""
    if now is None:
        now = datetime.now()
    sec = (now - start).total_seconds()
    if sec < 0:
        return "시계 보정 필요"
    if sec < 60:
        return "1분 미만"
    mins = int(sec // 60)
    if mins < 60:
        return f"{mins}분"
    h, m = divmod(mins, 60)
    if h < 24:
        return f"{h}시간 {m}분" if m else f"{h}시간"
    d, rh = divmod(h, 24)
    return f"{d}일 {rh}시간" if rh else f"{d}일"


@app.route("/api/qm-remote/config", methods=["GET"])
def api_qm_remote_config_get():
    cfg = _load_qm_remote_config()
    pc_id = _qm_client_pc_id()
    display_name = _qm_local_get_display_name(pc_id)
    tok = str(cfg.get("desktop_api_token") or "").strip()
    return jsonify(
        {
            "ok": True,
            "servers": cfg.get("servers") or [],
            "request_timeout_sec": float(cfg.get("request_timeout_sec") or 5),
            "my_pc_id": pc_id,
            "my_pc_display_name": display_name,
            "session_admin_configured": bool(_qm_effective_session_admin_pw(cfg.get("session_admin_password"))),
            "desktop_api_token_configured": bool(tok),
        }
    )


@app.route("/api/qm-remote/config", methods=["POST"])
def api_qm_remote_config_post():
    data = request.get_json(force=True, silent=True) or {}
    cur = _load_qm_remote_config()
    if "request_timeout_sec" in data:
        try:
            cur["request_timeout_sec"] = max(1.0, min(30.0, float(data["request_timeout_sec"])))
        except Exception:
            pass
    if "servers" in data:
        svl = data.get("servers")
        if not isinstance(svl, list):
            return jsonify({"ok": False, "error": "servers는 문자열 배열이어야 합니다"}), 400
        cur["servers"] = [str(x).strip() for x in svl if str(x).strip()]
    if not cur.get("servers"):
        cur["servers"] = list(_default_qm_remote_config()["servers"])
    if data.get("session_admin_password_clear") is True:
        cur["session_admin_password"] = QM_REMOTE_DEFAULT_SESSION_ADMIN_PW
    elif "session_admin_password" in data and str(data.get("session_admin_password") or "").strip():
        cur["session_admin_password"] = str(data["session_admin_password"]).strip()[:120]
    if data.get("desktop_api_token_clear") is True:
        cur["desktop_api_token"] = ""
    elif "desktop_api_token" in data and str(data.get("desktop_api_token") or "").strip():
        cur["desktop_api_token"] = str(data["desktop_api_token"]).strip()[:200]
    _save_qm_remote_config(cur)
    tok = str(cur.get("desktop_api_token") or "").strip()
    return jsonify(
        {
            "ok": True,
            "servers": cur["servers"],
            "request_timeout_sec": float(cur["request_timeout_sec"]),
            "session_admin_configured": bool(_qm_effective_session_admin_pw(cur.get("session_admin_password"))),
            "desktop_api_token_configured": bool(tok),
        }
    )


@app.route("/api/qm-remote/my-pc", methods=["POST"])
def api_qm_remote_my_pc_post():
    data = request.get_json(force=True, silent=True) or {}
    name = str(data.get("display_name") or data.get("name") or "").strip()[:200]
    pc_id = _qm_client_pc_id()
    _qm_local_set_display_name(pc_id, name)
    return jsonify({"ok": True, "my_pc_id": pc_id, "my_pc_display_name": name})


@app.route("/api/qm-remote/session/start", methods=["POST"])
def api_qm_remote_session_start():
    data = request.get_json(force=True, silent=True) or {}
    server_name = (data.get("server_name") or "").strip()
    if not server_name:
        return jsonify({"ok": False, "error": "server_name이 필요합니다"}), 400
    cfg = _load_qm_remote_config()
    allowed = {str(s).strip() for s in (cfg.get("servers") or []) if str(s).strip()}
    if server_name not in allowed:
        return jsonify({"ok": False, "error": "등록된 서버가 아닙니다"}), 400
    timeout = float(cfg.get("request_timeout_sec") or 5)
    timeout = max(1.0, min(30.0, timeout))
    pc_id = _qm_client_pc_id()
    me = _qm_session_effective_user(cfg, pc_id, timeout)

    cur = _qm_normalize_state(_qm_local_get_state(server_name))
    if cur["status"] == "ON":
        u = cur["user"]
        if u == me:
            return jsonify({"ok": True, "already": True, "server_name": server_name, "user": me})
        return (
            jsonify(
                {
                    "ok": False,
                    "error": f"이미 다른 사용자가 사용 중입니다: {u}",
                    "busy_user": u,
                }
            ),
            409,
        )
    busy_other = _qm_find_other_room_occupied_by_user(cfg.get("servers") or [], me, server_name)
    if busy_other:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": f"이미 「{busy_other}」에서 원격을 사용 중입니다. 먼저 그쪽에서 사용 종료 후 다시 시작하세요.",
                    "busy_room": busy_other,
                }
            ),
            409,
        )
    new_state = {
        "status": "ON",
        "user": me,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    _qm_local_set_state(server_name, new_state)
    return jsonify({"ok": True, "server_name": server_name, **new_state})


@app.route("/api/qm-remote/session/stop", methods=["POST"])
def api_qm_remote_session_stop():
    data = request.get_json(force=True, silent=True) or {}
    server_name = (data.get("server_name") or "").strip()
    admin_pw = str(data.get("admin_password") or "").strip()
    if not server_name:
        return jsonify({"ok": False, "error": "server_name이 필요합니다"}), 400
    cfg = _load_qm_remote_config()
    allowed = {str(s).strip() for s in (cfg.get("servers") or []) if str(s).strip()}
    if server_name not in allowed:
        return jsonify({"ok": False, "error": "등록된 서버가 아닙니다"}), 400
    timeout = float(cfg.get("request_timeout_sec") or 5)
    timeout = max(1.0, min(30.0, timeout))
    pc_id = _qm_client_pc_id()
    me = _qm_session_effective_user(cfg, pc_id, timeout)
    off_state = {"status": "OFF", "user": "", "timestamp": ""}

    cur = _qm_normalize_state(_qm_local_get_state(server_name))
    if cur["status"] != "ON":
        return jsonify({"ok": True, "already_off": True, "server_name": server_name})
    u = cur["user"]

    # 본인 세션이면 비밀번호 없이 즉시 종료
    if u == me:
        _qm_local_set_state(server_name, off_state)
        return jsonify({"ok": True, "server_name": server_name})

    # 타인 세션 — 비밀번호 없이 요청하면 입력 요청
    if not admin_pw:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "사용 종료를 위해 비밀번호를 입력하세요.",
                    "need_admin": True,
                    "busy_user": u,
                }
            ),
            403,
        )
    # 비밀번호 검증 — 일치하면 종료 허용
    stored = _qm_effective_session_admin_pw(cfg.get("session_admin_password"))
    try:
        pw_ok = len(admin_pw) == len(stored) and secrets.compare_digest(admin_pw, stored)
    except (TypeError, ValueError):
        pw_ok = False
    if not pw_ok:
        return (
            jsonify({"ok": False, "error": "비밀번호가 올바르지 않습니다."}),
            403,
        )
    _qm_local_set_state(server_name, off_state)
    return jsonify({"ok": True, "server_name": server_name})


def _qm_build_status_rows(servers: list, states: dict, notes: dict) -> list:
    rows = []
    for name in servers:
        if not isinstance(name, str) or not name.strip():
            continue
        name = name.strip()
        raw_st = states.get(name)
        if isinstance(raw_st, dict):
            st = _qm_normalize_state(raw_st)
        else:
            st = _qm_normalize_state({})
        status_on = st["status"] == "ON"
        user = st["user"]
        ts = st["timestamp"]
        note = ""
        if isinstance(notes, dict) and name in notes and notes[name] is not None:
            note = str(notes[name])[:2000]

        start_dt, start_label = _qm_parse_started_at(ts)
        if status_on:
            if start_dt:
                elapsed = _qm_format_elapsed(start_dt)
                start_display = start_label
            elif ts:
                start_display = ts
                elapsed = "(시각 형식 확인)"
            else:
                start_display = "—"
                elapsed = ""
        else:
            start_display = "—"
            elapsed = ""

        label = "사용 중" if status_on else "사용 가능"
        rows.append(
            {
                "server_name": name,
                "status": st["status"],
                "status_label": label,
                "user": user,
                "timestamp": ts,
                "start_display": start_display,
                "elapsed": elapsed,
                "note": note,
            }
        )
    return rows


@app.route("/api/qm-remote/status", methods=["GET"])
def api_qm_remote_status():
    cfg = _load_qm_remote_config()
    servers = cfg.get("servers") or []
    checked = datetime.now().isoformat(timespec="seconds")
    pc_id = _qm_client_pc_id()
    timeout = float(cfg.get("request_timeout_sec") or 5)
    my_user = _qm_session_effective_user(cfg, pc_id, timeout)
    try:
        slist = [str(s).strip() for s in servers if isinstance(s, str) and str(s).strip()]
        states = _qm_local_all_states(slist)
        notes = _qm_local_all_notes(slist)
        rows = _qm_build_status_rows(servers, states, notes)
        return jsonify(
            {
                "ok": True,
                "checked_at": checked,
                "storage": "local_sqlite",
                "fetch_error": "",
                "my_user": my_user,
                "rows": rows,
            }
        )
    except Exception as e:
        return jsonify(
            {
                "ok": False,
                "error": str(e)[:200],
                "checked_at": checked,
                "storage": "local_sqlite",
                "fetch_error": str(e)[:200],
                "my_user": my_user,
                "rows": [],
            }
        )


@app.route("/api/qm-remote/note", methods=["POST"])
def api_qm_remote_note():
    data = request.get_json(force=True, silent=True) or {}
    server_name = (data.get("server_name") or data.get("server") or "").strip()
    note = str(data.get("note") or "").strip()[:2000]
    if not server_name:
        return jsonify({"ok": False, "error": "서버명이 필요합니다"}), 400
    cfg = _load_qm_remote_config()
    allowed = {str(s).strip() for s in (cfg.get("servers") or []) if str(s).strip()}
    if server_name not in allowed:
        return jsonify({"ok": False, "error": "등록된 서버가 아닙니다"}), 400

    _qm_local_set_note(server_name, note)
    return jsonify({"ok": True, "server_name": server_name, "note": note})


# ─────────────────────────────────────────
#  웹 바로가기 (독립 설정: web_shortcuts.json)
# ─────────────────────────────────────────


def _default_web_shortcuts() -> dict:
    return {"links": []}


def _load_web_shortcuts() -> dict:
    if not WEB_SHORTCUTS_PATH.exists():
        data = _default_web_shortcuts()
        try:
            with open(WEB_SHORTCUTS_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        return data.copy()
    try:
        with open(WEB_SHORTCUTS_PATH, encoding="utf-8") as f:
            raw = json.load(f)
        if not isinstance(raw, dict):
            return _default_web_shortcuts()
        out = _default_web_shortcuts()
        out.update(raw)
        if not isinstance(out.get("links"), list):
            out["links"] = []
        return out
    except Exception:
        return _default_web_shortcuts()


def _save_web_shortcuts(data: dict) -> None:
    base = _default_web_shortcuts()
    base.update(data)
    if not isinstance(base.get("links"), list):
        base["links"] = []
    with open(WEB_SHORTCUTS_PATH, "w", encoding="utf-8") as f:
        json.dump(base, f, ensure_ascii=False, indent=2)


def _normalize_shortcut_url(url: str) -> str | None:
    u = (url or "").strip()
    if not u:
        return None
    low = u.lower()
    if low.startswith("javascript:") or low.startswith("data:") or low.startswith("vbscript:"):
        return None
    if low.startswith("http://") or low.startswith("https://"):
        return u[:2000]
    return None


def _shortcut_str_field(raw, max_len: int) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    return s[:max_len]


@app.route("/api/web-shortcuts", methods=["GET"])
def api_web_shortcuts_get():
    cfg = _load_web_shortcuts()
    links = cfg.get("links") or []
    clean_out = []
    if isinstance(links, list):
        for it in links:
            if not isinstance(it, dict):
                continue
            name = (it.get("name") or "").strip()
            url_ok = _normalize_shortcut_url(it.get("url") or "")
            if not name or not url_ok:
                continue
            tid = (it.get("id") or "").strip() or ("ls_" + str(int(time.time() * 1000)))
            pwd = it.get("password")
            if pwd is None and it.get("login_password") is not None:
                pwd = it.get("login_password")
            clean_out.append(
                {
                    "id": tid,
                    "name": name[:200],
                    "url": url_ok,
                    "login_id": _shortcut_str_field(it.get("login_id"), 200),
                    "password": _shortcut_str_field(pwd, 500),
                }
            )
    return jsonify({"ok": True, "links": clean_out})


@app.route("/api/web-shortcuts", methods=["POST"])
def api_web_shortcuts_post():
    data = request.get_json(force=True, silent=True) or {}
    links_in = data.get("links")
    if not isinstance(links_in, list):
        return jsonify({"ok": False, "error": "links는 배열이어야 합니다"}), 400
    clean = []
    for i, item in enumerate(links_in):
        if not isinstance(item, dict):
            continue
        name = (item.get("name") or "").strip()
        url_ok = _normalize_shortcut_url(item.get("url") or "")
        if not name or not url_ok:
            continue
        tid = (item.get("id") or "").strip()
        if not tid:
            tid = "ls_" + str(int(time.time() * 1000)) + "_" + str(i)
        pwd = item.get("password")
        if pwd is None and item.get("login_password") is not None:
            pwd = item.get("login_password")
        clean.append(
            {
                "id": tid,
                "name": name[:200],
                "url": url_ok,
                "login_id": _shortcut_str_field(item.get("login_id"), 200),
                "password": _shortcut_str_field(pwd, 500),
            }
        )
    _save_web_shortcuts({"links": clean})
    return jsonify({"ok": True, "links": clean})


# ─────────────────────────────────────────
#  API: 현황 리포트 — 서식 엑셀 (openpyxl)
# ─────────────────────────────────────────


def _xlsx_logger_summary(stats: object) -> str:
    if not isinstance(stats, dict):
        return "-"
    total = int(stats.get("total") or 0)
    if total <= 0:
        return "-"
    parts = []
    n = int(stats.get("normal") or 0)
    if n:
        parts.append(f"{n} 정상")
    n = int(stats.get("delayed") or 0)
    if n:
        parts.append(f"{n} 지연")
    n = int(stats.get("unconverted") or 0)
    if n:
        parts.append(f"{n} 미변환")
    n = int(stats.get("ghost") or 0)
    if n:
        parts.append(f"{n} Ghost")
    return ", ".join(parts) if parts else "-"


def _xlsx_manager_cell(site: dict) -> str:
    mgr = (site.get("manager") or "").strip()
    title = (site.get("manager_title") or "").strip()
    qm = (site.get("qm") or "").strip()
    if mgr:
        lines = [mgr]
        if title:
            lines.append(title)
        if qm:
            lines.append(f"QM {qm}")
        return "\n".join(lines)
    return "미배정"


@app.route("/api/report/export-xlsx", methods=["POST"])
def api_report_export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        return jsonify(
            {
                "ok": False,
                "error": "openpyxl 패키지가 필요합니다. pip install openpyxl",
            }
        ), 500

    data = request.get_json(force=True, silent=True) or {}
    groups = data.get("groups")
    if not isinstance(groups, list) or not groups:
        return jsonify({"ok": False, "error": "내보낼 데이터가 없습니다"}), 400

    title = (data.get("title") or "큐엠 자동화 관리 프로그램 — 현장 현황").strip()
    subtitle = (data.get("subtitle") or "").strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "현장현황"

    thin = Side(style="thin", color="FFCBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_banner = PatternFill(fill_type="solid", fgColor="FF2563EB")
    fill_head = PatternFill(fill_type="solid", fgColor="FFE2E8F0")
    fill_warn = PatternFill(fill_type="solid", fgColor="FFFFFBEB")
    fill_err = PatternFill(fill_type="solid", fgColor="FFFEF2F2")
    fill_white = PatternFill(fill_type="solid", fgColor="FFFFFFFF")

    font_banner = Font(name="맑은 고딕", size=13, bold=True, color="FFFFFFFF")
    font_title = Font(name="맑은 고딕", size=15, bold=True, color="FF0F172A")
    font_sub = Font(name="맑은 고딕", size=10, color="FF64748B")
    font_hdr = Font(name="맑은 고딕", size=10, bold=True, color="FF1D4ED8")
    font_body = Font(name="맑은 고딕", size=10, color="FF0F172A")

    al_top = Alignment(vertical="top", wrap_text=True)
    al_banner = Alignment(vertical="center", horizontal="left", indent=1)
    al_title = Alignment(vertical="center", horizontal="left")

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value=title)
    c.font = font_title
    c.alignment = al_title
    r += 1
    if subtitle:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(row=r, column=1, value=subtitle)
        c.font = font_sub
        c.alignment = al_title
        r += 1
    r += 1

    headers = ["현장명", "담당자", "로거 현황", "프로그램", "특이사항"]

    for g in groups:
        if not isinstance(g, dict):
            continue
        company = (g.get("company") or "").strip()
        sites = g.get("sites")
        if not isinstance(sites, list):
            sites = []
        n = len(sites)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(row=r, column=1, value=f"{company}  ·  {n}개 현장")
        c.fill = fill_banner
        c.font = font_banner
        c.alignment = al_banner
        r += 1

        for col, h in enumerate(headers, start=1):
            hc = ws.cell(row=r, column=col, value=h)
            hc.fill = fill_head
            hc.font = font_hdr
            hc.border = border
            hc.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        r += 1

        for site in sites:
            if not isinstance(site, dict):
                continue
            stats = site.get("stats") if isinstance(site.get("stats"), dict) else {}
            logger_txt = _xlsx_logger_summary(stats)
            mgr = _xlsx_manager_cell(site)
            prog = (site.get("program") or "").strip() or "-"
            memo = (site.get("memo") or "").strip() or "-"
            site_name = site.get("site") or ""

            row_fill = (
                fill_err
                if int(stats.get("unconverted") or 0) > 0
                else fill_warn
                if int(stats.get("delayed") or 0) > 0
                else fill_white
            )

            vals = [site_name, mgr, logger_txt, prog, memo]
            for col, val in enumerate(vals, start=1):
                dc = ws.cell(row=r, column=col, value=val)
                dc.font = font_body
                dc.fill = row_fill
                dc.border = border
                dc.alignment = al_top
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 40

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"cp3_status_{stamp}.xlsx",
    )


# ─────────────────────────────────────────
#  메인 페이지
# ─────────────────────────────────────────

@app.route("/")
def index():
    resp = app.make_response(render_template("index.html"))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


# ─────────────────────────────────────────
#  외부 통합용
# ─────────────────────────────────────────

_server_started = False
_server_lock = threading.Lock()


def start_server(open_browser: bool = False):
    global _server_started
    with _server_lock:
        if _server_started:
            return
        _server_started = True
    if open_browser:
        import webbrowser
        threading.Timer(1.5, lambda: webbrowser.open(f"http://localhost:{PORT}")).start()
    try:
        try:
            from waitress import serve
            print(f"[MonitoringServer] waitress 서버 시작 → 0.0.0.0:{PORT}", flush=True)
            serve(app, host="0.0.0.0", port=PORT, threads=8,
                  channel_timeout=120, cleanup_interval=30)
        except ImportError:
            print(f"[MonitoringServer] waitress 미설치 → Flask 개발서버 사용 (pip install waitress 권장)", flush=True)
            app.run(host="0.0.0.0", port=PORT, debug=False, use_reloader=False, threaded=True)
    except OSError as e:
        print(f"[MonitoringServer] 포트 {PORT} 바인딩 실패: {e}", file=sys.stderr)


if __name__ == "__main__":
    print(f"큐엠 자동화 관리 프로그램 모니터링 서버: http://localhost:{PORT}")
    start_server(open_browser=True)
